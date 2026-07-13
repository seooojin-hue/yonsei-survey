import uvicorn
import inspect
import traceback
import io
import urllib.parse
import pandas as pd
import shutil
from pydantic import BaseModel
from typing import List, Dict, Any
from contextlib import asynccontextmanager
from typing import Optional, Dict, Any
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import FastAPI, HTTPException, Body, Form, File, UploadFile, Response
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
import requests
from bs4 import BeautifulSoup
from fastapi.staticfiles import StaticFiles
import json

# 우리가 쪼개놓은 파일들 임포트
from config import *
from utils import *
from reports_logic import *

# 1. lifespan 함수를 '먼저' 정의합니다.
@asynccontextmanager
async def lifespan(app: FastAPI):
    # ⭐️ 안전장치: uploads 폴더가 없으면 파이썬이 스스로 빈 폴더를 만듭니다.
    safe_upload_dir = globals().get('UPLOAD_DIR', 'uploads')
    os.makedirs(safe_upload_dir, exist_ok=True)
    
    # [Startup] 서버 시작 시 실행
    schema_manager.load_schema()
    load_report_templates()
    yield
# 2. 정의된 lifespan을 사용하여 app 인스턴스를 '나중에' 만듭니다.
app = FastAPI(lifespan=lifespan)

# 3. 미들웨어 설정
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/uploads", StaticFiles(directory=globals().get('UPLOAD_DIR', 'uploads')), name="uploads")

# 1. 보고서 레지스트리 (보고서명: 함수명)
REPORT_LOGIC_REGISTRY = {
    "[표 1.3.1-1] 프로그램 책임자 현황": fetch_custom_1_3_1_1,
    "[표 1.3.2-1] 운영지원 인력 현황": fetch_custom_1_3_2_1,
    "[표 1.3.3-1] 프로그램 책임자 및 보건의료정보관리사 교수의 프로그램 운영 의사결정 참여 실적": fetch_custom_1_3_3_1,
    "[표 2.2-1] 이수영역별 2023학년도 교육과정표": fetch_curriculum_tables,
    "[표 2.2-2] 이수영역별 2024학년도 교육과정표": fetch_curriculum_tables,
    "[표 2.2-3] 이수영역별 2025학년도 교육과정표": fetch_curriculum_tables,
    "[표 2.2-4] 이수영역별 2026학년도 교육과정표": fetch_curriculum_tables,
    "[표 2.2.2-1] 프로그램 교육과정 선-후수 체계 (2023학년도 교육과정)": fetch_custom_pre_post_system,
    "[표 2.2.2-2] 프로그램 교육과정 선-후수 체계 (2024학년도 교육과정)": fetch_custom_pre_post_system,
    "[표 2.2.2-3] 프로그램 교육과정 선-후수 체계 (2025학년도 교육과정)": fetch_custom_pre_post_system,
    "[표 2.2.2-5] 프로그램 교육과정 이수체계 준수 실적": fetch_custom_2_2_2_5,
    "[표 2.3.1-1] 2023학년도~2025학년도 필수이수 교과목 편성표": fetch_custom_2_3_1_1,
    "[표 2.3.1-2] 2026학년도 필수이수 교과목 편성표 (2주기 평가·인증 기준 적용)": fetch_custom_2_3_1_2,
    "[표 2.3.1-5] 2023학년도 필수이수 교과목 운영실적": fetch_custom_2_3_1_5,
    "[표 2.3.1-6] 2024학년도 필수이수 교과목 운영실적": fetch_custom_2_3_1_6,
    "[표 2.3.1-7] 2025학년도 필수이수 교과목 운영실적": fetch_custom_2_3_1_7,
    "[표 2.3.1-8] 2026학년도 필수이수 교과목 운영실적": fetch_custom_2_3_1_8,
    "[표 2.3.2-1] 2024학년도~2025학년도 필수이수 교과목 학습내용 요약": fetch_custom_2_3_2_1,
    "[표 2.3.2-2] 2026학년도 필수이수 교과목 학습내용 요약": fetch_custom_2_3_2_2,
    "[표 2.3.2-3] 2023학년도 필수이수 교과목 학습내용 반영 운영실적": fetch_custom_2_3_2_3,
    "[표 2.3.2-4] 2024학년도 필수이수 교과목 학습내용 반영 운영실적": fetch_custom_2_3_2_4,
    "[표 2.3.2-5] 2025학년도 필수이수 교과목 학습내용 반영 운영실적": fetch_custom_2_3_2_5,
    "[표 2.3.2-6] 2026학년도 필수이수 교과목 학습내용 반영 운영실적": fetch_custom_2_3_2_6,
    "[표 2.4.1-1] 2023학년도 선택이수 교과목 특성화 영역 현황": fetch_custom_2_4_1_1,
    "[표 2.4.1-2] 2024학년도 선택이수 교과목 특성화 영역 현황": fetch_custom_2_4_1_2,
    "[표 2.4.1-3] 2025학년도 선택이수 교과목 특성화 영역 현황": fetch_custom_2_4_1_3,
    "[표 2.4.1-4] 2026학년도 선택이수 교과목 특성화 영역 현황 (2주기 평가·인증 기준 적용)": fetch_custom_2_4_1_4,
    "[표 2.4.2-1] 2023학년도 ~ 2025학년도 선택이수 교과목 학점 편성 현황": fetch_custom_2_4_2_1,
    "[표 2.4.2-2] 2026학년도 선택이수 교과목 학점 편성 현황 (2주기 평가·인증 기준 적용)": fetch_custom_2_4_2_2,
    "[표 2.4.2-3] 2023학년도 선택이수 교과목 운영실적": fetch_custom_2_4_2_3,
    "[표 2.4.2-4] 2024학년도 선택이수 교과목 운영실적": fetch_custom_2_4_2_4,
    "[표 2.4.2-5] 2025학년도 선택이수 교과목 운영실적": fetch_custom_2_4_2_5,
    "[표 2.4.2-6] 2026학년도 선택이수 교과목 운영실적": fetch_custom_2_4_2_6,
    "[표 2.5.1-1] 2023학년도 ~ 2025학년도 이론 교과목의 효과적 교수학습방법 적용 계획": fetch_custom_2_5_1_1,
    "[표 2.5.1-2] 2026학년도 이론 교과목의 효과적 교수학습방법 적용 계획 (2주기 평가·인증 기준 적용 계획)": fetch_custom_2_5_1_2,
    "[표 2.5.2-3] 2023학년도 다양한 평가방법 적용 실적": fetch_custom_2_5_2_3,
    "[표 2.5.2-4] 2024학년도 다양한 평가방법 적용 실적": fetch_custom_2_5_2_4,
    "[표 2.5.2-5] 2025학년도 다양한 평가방법 적용 실적": fetch_custom_2_5_2_5,
    "[표 2.5.2-6] 2026학년도 다양한 평가방법 적용 실적": fetch_custom_2_5_2_6,
    "[표 2.6.1-1] 학습성과 성취도 종합적 분석 실적": fetch_custom_2_6_1_1,
    "[표 2.6.2-1] 교과목의 지속적 질 개선 실적": fetch_custom_2_6_2_1,
    "[표 2.6.2-2] 지속적 질 개선 체계 검토내용의 학과/학부 교수와 공유 실적": fetch_custom_2_6_2_2,
    "[표 2.6.2-3] 성취도가 일정 수준 미만 대상의 개선 계획": fetch_custom_2_6_2_3,
    "[표 4.1.2-2] 전임교원 수업시수 현황": fetch_custom_4_1_2_2,
    "[표 4.2.1-1] 필수이수 교과목에 대한 담당교수의 전공일치 현황": fetch_custom_4_2_1_1,
    "[표 4.2.2-1] 필수 및 선택이수 교과목 전임교원 강의 학점 분담 비율": fetch_custom_4_2_2_1,
    "[표 4.2.2-2] 필수 및 선택이수 교과목별 전임/비전임 교수의 강의 시수 현황": fetch_custom_4_2_2_2,
    "[표 5.1.2-2] 필수이수 실습 교과목의 실습실 운영 현황": fetch_custom_5_1_2_2,
    "[표 5.1.2-3] 보건의료정보관리 실무, 보건의료 통계, 보건의료데이터 관리 교과목의 전산 실습실 운영 현황": fetch_custom_5_1_2_3,
    "[표 5.1.2-4] 질병 및 의료행위 분류, 의무기록정보 분석 실무, 의무기록정보 질 향상 실무, 건강보험 이론 및 실무, 암 등록 교과목의 실무/전산 실습실 운영 현황": fetch_custom_5_1_2_4,
    "표 I-3 교수진 강의담당 분석": fetch_custom_I_3,
    "[별책] 시간표": fetch_timetable_report
}

AUTO_REPORT_MAPPING = {
    "[표 5.1.1-1] 학과/학부의 확보 기준 면적 (제출일 기준)": "공간시설",
    "[표 5.1.2-1] 실습실 확보 현황 (제출일 기준)": "실습실현황",
}


@app.get("/api/db-list")
def get_db_list(): return {"dbs": schema_manager.get_db_list()}

@app.get("/api/schema/{db_name}")
def get_schema(db_name: str, year: str = None): 
    # 1. 파일 경로 설정
    filename = f"{year}_draft.csv" if year else "draft.csv"
    file_path = os.path.join(UPLOAD_DIR, db_name, filename)
    
    dynamic_cols = []
    # 2. 만능 비서 smart_read_df로 헤더(컬럼명)만 읽어오기
    if os.path.exists(file_path):
        df_h = smart_read_df(file_path, nrows=0) # 데이터 없이 컬럼명만 로드
        if df_h is not None:
            dynamic_cols = df_h.columns.tolist()

    # 3. schema_manager에서 정의된 컬럼 정보 가져오기
    cols = schema_manager.get_columns(db_name, dynamic_cols=dynamic_cols) or []
    
    # 4. 방어 로직: 한글 항목명(label)이 없으면 변수명(name)을 라벨로 사용
    for col in cols:
        if not col.get('label') or str(col.get('label')).strip() == "":
            col['label'] = col.get('name', 'Unknown')
            
    return {"db_name": db_name, "columns": cols}

@app.get("/api/versions/{db_name}")
def get_versions(db_name: str, year: str = None):
    save_dir = os.path.join(UPLOAD_DIR, db_name)
    if not os.path.exists(save_dir):
        return {"versions": []}
    
    # 파일 목록을 가져온 뒤 v로 시작하는 파일만 필터링
    files = [f for f in os.listdir(save_dir) if f.startswith('v') and f.endswith('.csv')]
    
    # ★ 교과목 DB처럼 연도별 분리가 필요한 경우 해당 연도 문자열이 포함된 파일만 필터링
    if year:
        files = [f for f in files if year in f]
    
    # 최신순 정렬
    files.sort(reverse=True)
    return {"versions": files}

@app.get("/api/load-version/{db_name}/{filename}")
def load_version(db_name: str, filename: str):
    file_path = os.path.join(UPLOAD_DIR, db_name, filename)
    df = smart_read_df(file_path)
    if df is None:
        raise HTTPException(status_code=404, detail="File not found")
    try:
        df = df.fillna("")
        
        df = enforce_year_sorting(df, db_name)
        
        return {"status": "success", "rows": df.to_dict(orient='records')}
        
    except Exception as e:
        print(f"❌ 버전 로드 중 오류 발생: {e}")
        raise HTTPException(status_code=500, detail="Error processing file")

@app.get("/api/load-draft/{db_name}")
def load_draft(db_name: str, year: Optional[str] = None):
    # 1. 특정 DB는 '최종 저장' 후 입력된 데이터가 안 보이도록 draft 파일만 별도 확인합니다.
    clear_on_save_dbs = ['교수인적사항', '프로그램 최종성과 평가모델', '운영지원인력', '학생', '학교현황']
    
    if db_name in clear_on_save_dbs:
        filename = f"{year}_draft.csv" if year else "draft.csv"
        file_path = os.path.join(UPLOAD_DIR, db_name, filename)
        
        # 최종 저장을 눌러 draft 파일이 삭제된 상태라면 빈 데이터를 반환하여 화면을 비웁니다.
        if not os.path.exists(file_path):
            return {"status": "success", "rows": [], "count": 0}
            
        df = smart_read_df(file_path)
        if df is None:
            return {"status": "success", "rows": [], "count": 0}
            
        df = df.fillna("")
        return {
            "status": "success", 
            "rows": df.to_dict(orient='records'),
            "count": len(df)
        }

    # 2. 그 외 DB는 기존처럼 병합된 전체 데이터를 불러옵니다.
    df, error = get_db_dataframe(db_name)
    
    if error or df is None:
        return {"status": "no_draft", "rows": [], "message": error}

    if year:
        year_str = str(year).strip()
        year_col = next((c for c in df.columns if c in ['구분', 'curr_year', '연도', '학년도']), None)
        
        if year_col:
            df = df[df[year_col].astype(str).str.contains(year_str)]
        else:
            return {"status": "success", "rows": [], "count": 0}

    df = df.fillna("")

    return {
        "status": "success", 
        "rows": df.to_dict(orient='records'),
        "count": len(df)
    }

@app.post("/api/save/temp/{db_name}")
def save_temp(db_name: str, payload: Dict[str, Any] = Body(...)):
    try:
        rows = payload.get("data", [])
        year = payload.get("year")
        year_str = str(year) if year else ""
        
        save_dir = os.path.join(UPLOAD_DIR, db_name)
        os.makedirs(save_dir, exist_ok=True)
        
        df = pd.DataFrame(rows)
        if not df.empty:
            cols_to_drop = [c for c in ['rowKey', '_attributes'] if c in df.columns]
            if cols_to_drop:
                df = df.drop(columns=cols_to_drop)
                
        filename = f"{year_str}_draft.csv" if year_str else "draft.csv"
        file_path = os.path.join(save_dir, filename)
        
        df.to_csv(file_path, index=False, encoding='utf-8-sig')
        return {"status": "success"}
        
    except PermissionError:
        raise HTTPException(status_code=400, detail="파일이 현재 다른 프로그램(엑셀 등)에서 열려 있습니다. 파일을 완전히 닫고 다시 시도해주세요.")
    except Exception as e:
        print(f"❌ 임시저장 에러: {e}")
        raise HTTPException(status_code=500, detail=f"서버 내부 오류: {str(e)}")

@app.post("/api/save/final/{db_name}")
def save_final(db_name: str, payload: Dict[str, Any] = Body(...)):
    try:
        rows = payload.get("data", [])
        year = payload.get("year")
        year_str = str(year) if year else ""
        
        save_dir = os.path.join(UPLOAD_DIR, db_name)
        os.makedirs(save_dir, exist_ok=True)
        
        df = pd.DataFrame(rows)
        if not df.empty:
            cols_to_drop = [c for c in ['rowKey', '_attributes'] if c in df.columns]
            if cols_to_drop:
                df = df.drop(columns=cols_to_drop)
                
        date_str = datetime.datetime.now().strftime('%Y%m%d')
        
        version = 1
        for f in os.listdir(save_dir):
            if f.startswith('v') and f.endswith('.csv'):
                if year_str:
                    if year_str in f:
                        version += 1
                else:
                    version += 1
                    
        fname = f"v{version}_{year_str + '_' if year_str else ''}{date_str}.csv"
        file_path = os.path.join(save_dir, fname)
        
        df.to_csv(file_path, index=False, encoding='utf-8-sig')
        
        # 임시 파일 삭제 (삭제 실패하더라도 저장은 성공하도록 예외 처리)
        draft_filename = f"{year_str}_draft.csv" if year_str else "draft.csv"
        draft_path = os.path.join(save_dir, draft_filename)
        if os.path.exists(draft_path): 
            try:
                os.remove(draft_path)
            except Exception:
                pass 
            
        return {"status": "success", "filename": fname}
        
    except PermissionError:
        raise HTTPException(status_code=400, detail="파일이 현재 다른 프로그램(엑셀 등)에서 열려 있습니다. 파일을 완전히 닫고 다시 시도해주세요.")
    except Exception as e:
        print(f"❌ 최종저장 에러: {e}")
        raise HTTPException(status_code=500, detail=f"서버 내부 오류: {str(e)}")
    
@app.post("/api/upload/{db_name}")
async def upload_file(
    db_name: str, 
    files: List[UploadFile] = File(...), 
    year: str = Form(None)
):
    save_dir = os.path.join(UPLOAD_DIR, db_name)
    os.makedirs(save_dir, exist_ok=True)
    
    filename = f"{year}_draft.csv" if year else "draft.csv"
    draft_path = os.path.join(save_dir, filename)
    all_dfs = []
    
    try:
        for file in files:
            temp_path = os.path.join(save_dir, f"temp_{file.filename}")
            with open(temp_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            # 일반 DB 업로드 파서
            if file.filename.lower().endswith(('.xlsx', '.xls')):
                xls = pd.read_excel(temp_path, sheet_name=None, engine='openpyxl')
                sheet_dfs = [df for df in xls.values() if not df.empty]
                if sheet_dfs: all_dfs.append(pd.concat(sheet_dfs, ignore_index=True))
            else:
                try: all_dfs.append(pd.read_csv(temp_path, encoding='cp949'))
                except: all_dfs.append(pd.read_csv(temp_path, encoding='utf-8-sig'))
                    
            if os.path.exists(temp_path): os.remove(temp_path)

        if not all_dfs:
            raise Exception("유효한 데이터가 없습니다.")

        combined_df = pd.concat(all_dfs, ignore_index=True)

        # 컬럼 매핑 및 정제
        combined_df = normalize_headers_with_alias(combined_df)
        schema_cols = schema_manager.get_columns(db_name)
        mapping = {col['label'].strip(): col['name'] for col in schema_cols}
        combined_df.rename(columns=mapping, inplace=True)
        
        valid_cols = [col['name'] for col in schema_cols]
        if valid_cols:
            combined_df = combined_df[[c for c in valid_cols if c in combined_df.columns]]

        if year:
            target_year_col = next((c for c in combined_df.columns if any(k in c for k in ['연도', 'curr', 'year'])), '연도')
            combined_df[target_year_col] = year

        combined_df = combined_df.where(pd.notnull(combined_df), "")
        combined_df.to_csv(draft_path, index=False, encoding='utf-8-sig')
        
        return {"status": "success", "rows": combined_df.to_dict(orient='records')}
        
    except Exception as e:
        print(f"🔥 파일 업로드 에러: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/api/reports/list')
def get_report_list():
    path_xlsx = os.path.join(BASE_DIR, "report_list.xlsx")
    path_csv = os.path.join(BASE_DIR, "report_list.csv")
    df = smart_read_df(path_xlsx)
    if df is None:
        df = smart_read_df(path_csv)
    if df is None: 
        return []
    try:
        reports = []
        for _, row in df.iterrows():
            title = str(row.get('표이름', '')).strip()
            status = str(row.get('진행사항', 'X')).strip()
            if not title or title.lower() == 'nan': 
                continue
            reports.append({"id": title, "title": title, "status": status})
        reports.sort(key=lambda x: (
            0 if x['title'] in REPORT_LOGIC_REGISTRY else 1, 
            x['title']
        ))
        return reports
    except Exception as e: 
        print(f"❌ 보고서 목록 생성 중 오류 발생: {e}")
        return []

def fetch_data_generic(report_id, db_name):
    clean_title = normalize_title(report_id)
    template_headers = REPORT_TEMPLATES.get(clean_title)
    df, err = get_db_dataframe(db_name)
    if df is None: return {"headers": template_headers or [], "rows": [], "message": err}
    if template_headers:
        final_rows = apply_aliases_and_template(df, template_headers, db_name)
        return {"headers": template_headers, "rows": final_rows}
    else:
        schema_cols = schema_manager.get_columns(db_name)
        if schema_cols:
            reverse_map = {col['name']: col['label'] for col in schema_cols}
            df.rename(columns=reverse_map, inplace=True)
        return {"headers": df.columns.tolist(), "rows": df.to_dict(orient='records')}

@app.get('/api/report/preview')
def preview_report(report_id: str):
    clean_req = normalize_title(report_id)
    print(f"\n--- [진단 시작] 보고서명: {report_id} ---")
    
    # 1. 레지스트리 매칭 확인
    func = None
    for key, f in REPORT_LOGIC_REGISTRY.items():
        if normalize_title(key) == clean_req:
            func = f
            break
    
    if not func:
        print(f"❌ 오류: 레지스트리에 '{report_id}'와 매칭되는 함수가 없습니다.")
        return {"headers": [], "rows": [], "message": "로직 매칭 실패"}

    # 2. 함수 실행 및 내부 에러 추적
    try:
        sig = inspect.signature(func)
        if 'report_id' in sig.parameters:
            result = func(report_id)
        else:
            result = func()
            
        # 3. 결과 데이터 검증
        if not result.get("rows"):
            print(f"⚠️ 경고: 함수 '{func.__name__}'가 실행되었으나 빈 데이터를 반환했습니다. (DB 파일이나 연도 필터링 확인 필요)")
        else:
            print(f"✅ 성공: {len(result['rows'])}개의 행을 불러왔습니다.")
            
        return result

    except Exception as e:
        print(f"🔥 치명적 에러: {func.__name__} 실행 중 오류 발생 -> {str(e)}")
        # 에러 발생 지점(Line number) 출력
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/api/report/download/excel')
def download_report_excel(report_id: str):
    try:
        clean_req = normalize_title(report_id)
        data = None
        
        # [기존 로직] 보고서 데이터를 가져옵니다.
        for key, func in REPORT_LOGIC_REGISTRY.items():
            if normalize_title(key) == clean_req:
                sig = inspect.signature(func)
                if 'report_id' in sig.parameters:
                    data = func(report_id)
                else:
                    data = func()
                break
        
        if not data:
            for key, val in AUTO_REPORT_MAPPING.items():
                if normalize_title(key) == clean_req:
                    data = fetch_data_generic(report_id, val)
                    break
                
        if not data:
            raise HTTPException(status_code=404, detail="No data found")
        if data.get("message") and not data.get("rows"):
            raise HTTPException(status_code=400, detail=data["message"])
        
        # 1. 데이터프레임 변환 (화면과 동일한 열 순서 보장)
        headers = data.get('headers', list(data['rows'][0].keys()) if data['rows'] else [])
        df = pd.DataFrame(data['rows'], columns=headers)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            safe_sheet_name = report_id.replace('[', '').replace(']', '').strip()[:31]
            
            # ====================================================
            # ⭐️ [표 2.3.1-2] 또는 [표 2.4.2-2] (2026학년도 전용) 셀 병합 로직
            # ====================================================
            if '2.3.1-2' in report_id or '2.4.2-2' in report_id:
                # 데이터만 3번째 줄(startrow=2)부터 기록 (1, 2번째 줄은 우리가 직접 그림)
                df.to_excel(writer, index=False, header=False, startrow=2, sheet_name=safe_sheet_name)
                ws = writer.sheets[safe_sheet_name]
                
                # 1행 다중 헤더 작성
                ws['A1'] = '교과목명'
                ws['B1'] = '2026학년도'
                ws['E1'] = '변경 내용'
                
                # 2행 서브 헤더 작성
                ws['B2'] = '개설학년/학기'
                ws['C2'] = '학점'
                ws['D2'] = '실습 여부'
                
                # 셀 병합 (RowSpan, ColSpan)
                ws.merge_cells('A1:A2') # 교과목명 세로 2칸 병합
                ws.merge_cells('B1:D1') # 2026학년도 가로 3칸 병합
                ws.merge_cells('E1:E2') # 변경 내용 세로 2칸 병합
                
                # 스타일 지정 (배경색: 회색, 굵게, 가운데 정렬)
                header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                header_font = Font(bold=True)
                align_center = Alignment(horizontal='center', vertical='center')
                
                # 헤더(1~2행)에 스타일 적용
                for r in range(1, 3):
                    for c in range(1, 6):
                        cell = ws.cell(row=r, column=c)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = align_center
                
                # 데이터 영역(3행부터 끝까지)에 가운데 정렬 적용
                for r in range(3, ws.max_row + 1):
                    for c in range(1, 6):
                        ws.cell(row=r, column=c).alignment = align_center
                        
                # 엑셀 열(Column) 너비 보기 좋게 조절
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 10
                ws.column_dimensions['D'].width = 10
                ws.column_dimensions['E'].width = 20

            # ====================================================
            # ⭐️ [표 2.3.1-1] 또는 [표 2.4.2-1] (2023~2025학년도 3개년)
            # ====================================================
            elif '2.3.1-1' in report_id or '2.4.2-1' in report_id:
                df.to_excel(writer, index=False, header=False, startrow=2, sheet_name=safe_sheet_name)
                ws = writer.sheets[safe_sheet_name]
                
                ws['A1'] = '교과목명'
                ws['B1'] = '2023학년도'
                ws['E1'] = '2024학년도'
                ws['H1'] = '2025학년도'
                ws['K1'] = '변경 내용'
                
                sub_headers = ['개설학년/학기', '학점', '실습 여부']
                for i in range(3):
                    ws.cell(row=2, column=2+i, value=sub_headers[i])
                    ws.cell(row=2, column=5+i, value=sub_headers[i])
                    ws.cell(row=2, column=8+i, value=sub_headers[i])
                    
                ws.merge_cells('A1:A2')
                ws.merge_cells('B1:D1')
                ws.merge_cells('E1:G1')
                ws.merge_cells('H1:J1')
                ws.merge_cells('K1:K2')
                
                header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                header_font = Font(bold=True)
                align_center = Alignment(horizontal='center', vertical='center')
                
                for r in range(1, 3):
                    for c in range(1, 12):
                        cell = ws.cell(row=r, column=c)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = align_center
                        
                for r in range(3, ws.max_row + 1):
                    for c in range(1, 12):
                        ws.cell(row=r, column=c).alignment = align_center
                        
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['K'].width = 15
                
            # ====================================================
            # 기본 표 (복잡한 병합이 필요 없는 나머지 표들)
            # ====================================================
            else:
                df.to_excel(writer, index=False, sheet_name=safe_sheet_name)
        
        output.seek(0)
        encoded_filename = urllib.parse.quote(f"{report_id}.xlsx")
        
        return Response(
            content=output.getvalue(), 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
        
    except Exception as e:
        print(f"보고서 엑셀 다운로드 에러: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/reset/{db_name}")
def reset_db(db_name: str):
    save_dir = os.path.join(UPLOAD_DIR, db_name)
    if os.path.exists(save_dir):
        draft_path = os.path.join(save_dir, "draft.csv")
        if os.path.exists(draft_path):
            os.remove(draft_path) # 파일 삭제
            return {"status": "success", "message": "초기화 완료"}
    return {"status": "error", "message": "삭제할 데이터가 없습니다."}

# ==========================================
# ★★★ [성적 및 PO 성취도 분석 전용 API] ★★★
# ==========================================

class GradeData(BaseModel):
    year: str
    data: List[dict]

@app.post("/api/save/final/성적")
async def save_grades(grade_data: GradeData):
    try:
        save_dir = os.path.join(globals().get('UPLOAD_DIR', 'uploads'), "성적")
        os.makedirs(save_dir, exist_ok=True)
        
        file_path = os.path.join(save_dir, f"grades_{grade_data.year}.json")
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(grade_data.data, f, ensure_ascii=False, indent=2)
        
        return {"success": True, "message": "성적 데이터 최종 저장 완료"}
    except Exception as e:
        print(f"❌ 성적 저장 중 오류 발생: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/grades/{year}")
async def load_grades(year: str):
    try:
        save_dir = os.path.join(globals().get('UPLOAD_DIR', 'uploads'), "성적")
        file_path = os.path.join(save_dir, f"grades_{year}.json")
        
        if not os.path.exists(file_path):
            return []
        
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        return data
    except Exception as e:
        print(f"❌ 성적 로드 중 오류 발생: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# ★★★ [명세서 DB 전용 API 7개] ★★★
# ==========================================

@app.get("/api/명세서/courses")
def get_myungseseo_courses(year: str):
    df, err = get_db_dataframe("교과목")
    if df is None: return {"courses": []}
    
    df.columns = [str(c).strip() for c in df.columns]
    
    name_col = next((c for c in df.columns if re.sub(r'\s+', '', c) in ['교과목명', '과목명', '수업명', 'subject', '강좌명', 'course_name']), None)
    year_col = next((c for c in df.columns if re.sub(r'\s+', '', c) in ['구분', '학년도', '연도', 'year', 'curr', 'curr_year']), None)
    sem_col = next((c for c in df.columns if re.sub(r'\s+', '', c) in ['개설학년학기', '학년-학기', '개설학기', 'open_sem']), None)
    area_col = next((c for c in df.columns if re.sub(r'\s+', '', c) in ['이수구분', '필수선택', 'area_1']), None)
    
    if not all([name_col, year_col, sem_col, area_col]):
        return {"courses": []}
        
    res = []
    for _, row in df.iterrows():
        c_year = str(row.get(year_col, '')).strip()
        c_area = str(row.get(area_col, '')).strip()
        if c_year != year or c_area not in ['필수이수', '선택이수']: continue
        
        sem_raw = str(row.get(sem_col, '')).strip()
        nums = re.findall(r'\d+', sem_raw)
        g_val, s_val = (int(nums[0]), int(nums[1])) if len(nums) >= 2 else (1, 1) if len(nums) == 1 else (0, 0)
        
        actual = f"{int(year) + g_val - 1}-{s_val}학기" if g_val else ""
        
        res.append({
            "grade_sem": sem_raw,
            "actual_sem": actual,
            "course_name": str(row.get(name_col, '')).strip(),
            "area_1": c_area
        })
        
    # ==========================================
    # ★ 맞춤형 정렬 로직 (필수이수 지정 순서 -> 선택이수 학년학기순)
    # ==========================================
    def sort_key(item):
        area = item["area_1"]
        c_name = item["course_name"].replace(" ", "")
        g_sem = item["grade_sem"]
        
        # 1순위: 필수이수(0)가 선택이수(1)보다 무조건 위로
        area_order = 0 if area == '필수이수' else 1
        
        # 2순위: 필수이수 18개 과목 지정 순서 매핑
        custom_idx = 999
        if area_order == 0:
            if "보건의료정보관리학" in c_name: custom_idx = 1
            elif "보건의료정보관리실무" in c_name: custom_idx = 2
            elif "보건의료조직" in c_name: custom_idx = 3
            elif "건강정보보호" in c_name: custom_idx = 4
            elif "질병" in c_name and "분류" in c_name: custom_idx = 5
            elif "의무기록정보분석" in c_name: custom_idx = 6
            elif "의무기록정보질" in c_name: custom_idx = 7
            elif "암등록" in c_name: custom_idx = 8
            elif "의료의질" in c_name or "의료질" in c_name: custom_idx = 9
            elif "건강보험" in c_name: custom_idx = 10
            elif "보건의료통계" in c_name: custom_idx = 11
            elif "보건의료데이터" in c_name: custom_idx = 12
            elif "의료정보기술" in c_name: custom_idx = 13
            elif "의료관계법규" in c_name: custom_idx = 14
            elif "의학용어" in c_name: custom_idx = 15
            elif "병리학" in c_name: custom_idx = 16
            elif "해부생리" in c_name: custom_idx = 17
            elif "현장실습" in c_name: custom_idx = 18
            
        # 3순위: 학년-학기 (예: "1-1" < "2-2")
        # 4순위: 동점일 경우 이름순 (예: 의학용어(1), 의학용어(2))
        return (area_order, custom_idx, g_sem, item["course_name"])
        
    res.sort(key=sort_key)
    return {"courses": res}

@app.post("/api/명세서/save-temp")
def save_myungseseo_temp(payload: dict = Body(...)):
    year = payload.get("year")
    cname = payload.get("course_name")
    data = payload.get("data", {})
    
    s_dir = os.path.join(UPLOAD_DIR, "명세서", str(year))
    os.makedirs(s_dir, exist_ok=True)
    
    df = pd.DataFrame([data])
    df = df.where(pd.notnull(df), "")
    df.to_csv(os.path.join(s_dir, f"{cname}_draft.csv"), index=False, encoding='utf-8-sig')
    return {"status": "success"}

@app.post("/api/명세서/save-final")
def save_myungseseo_final(payload: dict = Body(...)):
    year = payload.get("year")
    cname = payload.get("course_name")
    data = payload.get("data", {})
    
    s_dir = os.path.join(UPLOAD_DIR, "명세서", str(year))
    os.makedirs(s_dir, exist_ok=True)
    
    ver = 1
    for f in os.listdir(s_dir):
        if f.endswith('.csv') and f.startswith('v') and cname in f: ver += 1
            
    fname = f"v{ver}_{cname}_{datetime.datetime.now().strftime('%Y%m%d')}.csv"
    
    df = pd.DataFrame([data])
    df = df.where(pd.notnull(df), "")
    df.to_csv(os.path.join(s_dir, fname), index=False, encoding='utf-8-sig')
    
    draft = os.path.join(s_dir, f"{cname}_draft.csv")
    if os.path.exists(draft): os.remove(draft)
        
    return {"status": "success", "filename": fname}

@app.get("/api/명세서/load")
def load_myungseseo(year: str, course_name: str):
    s_dir = os.path.join(UPLOAD_DIR, "명세서", str(year))
    if not os.path.exists(s_dir): 
        return {"status": "no_data", "data": {}}
    target = os.path.join(s_dir, f"{course_name}_draft.csv")
    if not os.path.exists(target):
        files = [f for f in os.listdir(s_dir) if f.startswith('v') and course_name in f and f.endswith('.csv')]
        if not files: 
            return {"status": "no_data", "data": {}}
        files.sort(reverse=True)
        target = os.path.join(s_dir, files[0])
    df = smart_read_df(target)
    if df is None or df.empty:
        return {"status": "no_data", "data": {}}
    df = df.fillna("")
    return {"status": "success", "data": df.iloc[0].to_dict()}

@app.get("/api/명세서/copy-candidates")
def get_copy_candidates(course_name: str, current_year: str):
    base_dir = os.path.join(UPLOAD_DIR, "명세서")
    if not os.path.exists(base_dir): return {"candidates": []}
    
    cands = []
    for d in os.listdir(base_dir):
        if d == current_year or not os.path.isdir(os.path.join(base_dir, d)): continue
        files = [f for f in os.listdir(os.path.join(base_dir, d)) if f.startswith('v') and course_name in f and f.endswith('.csv')]
        files.sort(reverse=True)
        if files: cands.append({"year": d, "filename": files[0], "path": f"{d}/{files[0]}"})
            
    cands.sort(key=lambda x: x["year"], reverse=True)
    return {"candidates": cands}

@app.get("/api/명세서/load-for-copy")
def load_for_copy(path: str):
    safe_path = path.replace("..", "")
    target = os.path.join(UPLOAD_DIR, "명세서", safe_path)
    df = smart_read_df(target)
    if df is None:
        return {"status": "error", "message": "파일을 불러올 수 없습니다."}
    df = df.fillna("")
    data = df.iloc[0].to_dict() if not df.empty else {}
    
    return {"status": "success", "data": data}

@app.post("/api/명세서/download-excel")
def download_myungseseo_excel(payload: dict = Body(...)):
    try:
        from openpyxl.styles import Border, Side
        
        data = payload.get("data", {})
        wb = Workbook()
        ws = wb.active
        
        year_str = str(data.get('year', ''))
        cname = data.get('course_name', '명세서')
        
        # ==========================================
        # ★ 추가됨: 현재 교과목이 몇 번째인지(Index) 알아내기
        # ==========================================
        courses_res = get_myungseseo_courses(year_str)
        courses_list = courses_res.get("courses", [])
        c_idx = 1
        for i, c in enumerate(courses_list):
            if c["course_name"] == cname:
                c_idx = i + 1
                break
        
        grey_bg = PatternFill("solid", fgColor="D9D9D9")
        white_bg = PatternFill("solid", fgColor="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        def set_cell(r, c, val, fill=None, font_color="000000", bold=False):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if fill: cell.fill = fill
            cell.font = Font(color=font_color, bold=bold)
            return cell

        # ★ 제목 행에 번호 추가
        title_str = f"{c_idx}. {cname} (입학년도: {year_str})"
        ws.merge_cells("A1:L1")
        set_cell(1, 1, title_str, None, "000000", True)
        
        ws.merge_cells("A2:B2"); set_cell(2, 1, "교과목명", grey_bg, bold=True)
        ws.merge_cells("C2:D2"); set_cell(2, 3, cname, white_bg)
        ws.merge_cells("E2:F2"); set_cell(2, 5, "구분", grey_bg, bold=True)
        ws.merge_cells("G2:H2"); set_cell(2, 7, data.get('area_1', ''), white_bg)
        ws.merge_cells("I2:J2"); set_cell(2, 9, "책임교수", grey_bg, bold=True)
        ws.merge_cells("K2:L2"); set_cell(2, 11, data.get('professor', ''), white_bg)
        
        ws.merge_cells("A3:B3"); set_cell(3, 1, "교과목개요", grey_bg, bold=True)
        ws.merge_cells("C3:L3"); set_cell(3, 3, data.get('overview', ''), white_bg)
        
        ws.merge_cells("A4:B5"); set_cell(4, 1, "프로그램\n최종성과", grey_bg, bold=True)
        for i in range(10): set_cell(4, i+3, f"PO{i+1}", grey_bg, bold=True)
        for i in range(10): set_cell(5, i+3, data.get(f'po{i+1}', ''), white_bg, bold=True)
            
        try: clo_list = json.loads(data.get('clo_list', '[]'))
        except: clo_list = []
        try: eval_data = json.loads(data.get('eval_data', '[]'))
        except: eval_data = []
        
        r_idx = 6
        if len(clo_list) > 0: ws.merge_cells(f"A{r_idx}:B{r_idx + len(clo_list)}")
        else: ws.merge_cells(f"A{r_idx}:B{r_idx}")
            
        set_cell(r_idx, 1, "학습목표\n(교과목 학습성과)", grey_bg, bold=True)
        ws.merge_cells(f"C{r_idx}:H{r_idx}"); set_cell(r_idx, 3, "학습목표\n(교과목 학습성과 CLO)", grey_bg, bold=True)
        ws.merge_cells(f"I{r_idx}:J{r_idx}"); set_cell(r_idx, 9, "연계 PO", grey_bg, bold=True)
        ws.merge_cells(f"K{r_idx}:L{r_idx}"); set_cell(r_idx, 11, "PO 성취도평가\n반영 여부", grey_bg, bold=True)
        r_idx += 1
        
        for clo in clo_list:
            set_cell(r_idx, 3, clo.get('no', ''), grey_bg, bold=True)
            ws.merge_cells(f"D{r_idx}:H{r_idx}"); set_cell(r_idx, 4, clo.get('content', ''), white_bg)
            ws.merge_cells(f"I{r_idx}:J{r_idx}"); set_cell(r_idx, 9, data.get('linked_po',''), white_bg)
            ws.merge_cells(f"K{r_idx}:L{r_idx}"); set_cell(r_idx, 11, data.get('po_reflected',''), white_bg)
            r_idx += 1
            
        ws.merge_cells(f"A{r_idx}:B{r_idx}"); set_cell(r_idx, 1, "학습내용", grey_bg, bold=True)
        ws.merge_cells(f"C{r_idx}:L{r_idx}"); set_cell(r_idx, 3, data.get('learning_content', ''), white_bg)
        r_idx += 1
        
        ws.merge_cells(f"A{r_idx}:B{r_idx}"); set_cell(r_idx, 1, "교수법", grey_bg, bold=True)
        ws.merge_cells(f"C{r_idx}:L{r_idx}"); set_cell(r_idx, 3, data.get('teaching_method', ''), white_bg)
        r_idx += 1
        
        ecols = len(eval_data) if eval_data else 1
        def get_col_span(index):
            base = 10 // ecols
            rem = 10 % ecols
            return base + (1 if index < rem else 0)

        ws.merge_cells(f"A{r_idx}:A{r_idx+4}"); set_cell(r_idx, 1, "평가", grey_bg, bold=True)
        ws.merge_cells(f"B{r_idx}:B{r_idx+1}"); set_cell(r_idx, 2, "영역", grey_bg, bold=True)
        
        ws.merge_cells(start_row=r_idx, start_column=3, end_row=r_idx, end_column=12)
        set_cell(r_idx, 3, "학습목표(교과목 학습성과)", grey_bg, bold=True)
        r_idx += 1
        
        c_start = 3
        for i in range(ecols):
            span = get_col_span(i)
            c_end = c_start + span - 1
            if span > 1: ws.merge_cells(start_row=r_idx, start_column=c_start, end_row=r_idx, end_column=c_end)
            set_cell(r_idx, c_start, f"CLO{i+1}", grey_bg, bold=True)
            c_start = c_end + 1
        r_idx += 1
        
        set_cell(r_idx, 2, "평가기준(배점율)", grey_bg, bold=True)
        c_start = 3
        for i, ev in enumerate(eval_data):
            span = get_col_span(i)
            c_end = c_start + span - 1
            if span > 1: ws.merge_cells(start_row=r_idx, start_column=c_start, end_row=r_idx, end_column=c_end)
            set_cell(r_idx, c_start, ev.get('criteria',''), white_bg)
            c_start = c_end + 1
        r_idx += 1
        
        set_cell(r_idx, 2, "평가방법", grey_bg, bold=True)
        c_start = 3
        for i, ev in enumerate(eval_data):
            span = get_col_span(i)
            c_end = c_start + span - 1
            if span > 1: ws.merge_cells(start_row=r_idx, start_column=c_start, end_row=r_idx, end_column=c_end)
            set_cell(r_idx, c_start, ev.get('method',''), white_bg)
            c_start = c_end + 1
        r_idx += 1
        
        set_cell(r_idx, 2, "목표수준", grey_bg, bold=True)
        c_start = 3
        for i, ev in enumerate(eval_data):
            span = get_col_span(i)
            c_end = c_start + span - 1
            if span > 1: ws.merge_cells(start_row=r_idx, start_column=c_start, end_row=r_idx, end_column=c_end)
            set_cell(r_idx, c_start, ev.get('target',''), white_bg)
            c_start = c_end + 1
        r_idx += 1
        
        ws.merge_cells(f"A{r_idx}:B{r_idx}"); set_cell(r_idx, 1, "선수과목", grey_bg, bold=True)
        ws.merge_cells(f"C{r_idx}:L{r_idx}"); set_cell(r_idx, 3, data.get('prerequisite', ''), white_bg)
        
        for row in ws.iter_rows(min_row=1, max_row=r_idx, min_col=1, max_col=12):
            for cell in row:
                cell.border = thin_border

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        # ★ 파일명에 번호 추가
        encoded_filename = urllib.parse.quote(f"{title_str}.xlsx")
        
        return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"})
    except Exception as e:
        print("엑셀 다운로드 에러:", str(e))
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/명세서/download-all-excel")
def download_all_myungseseo_excel(year: str):
    """해당 학년도의 모든 교과목 명세서를 하나의 엑셀 파일로 병합하여 다운로드합니다."""
    try:
        # 1. 교과목 DB 로드 및 대상 과목 추출
        df_courses, _ = get_db_dataframe("교과목")
        if df_courses is None:
            raise Exception("교과목 DB를 불러올 수 없습니다.")
            
        # find_col 도구를 사용하여 유연하게 컬럼 식별
        name_col = find_col(df_courses, ['교과목명', '과목명', 'course_name'])
        year_col = find_col(df_courses, ['구분', '학년도', '연도', 'curr_year'])
        sem_col  = find_col(df_courses, ['개설학년학기', '학년-학기', 'open_sem'])
        area_col = find_col(df_courses, ['이수구분', 'area_1', '학교이수구분'])
        
        if not all([name_col, year_col, sem_col, area_col]):
            raise Exception("필수 컬럼을 찾을 수 없습니다. DB 설정을 확인하세요.")

        # 해당 연도의 필수/선택이수 과목 목록 생성
        courses_list = []
        for _, row in df_courses.iterrows():
            c_year = str(row.get(year_col, '')).strip()
            c_area = str(row.get(area_col, '')).strip()
            if c_year != year or c_area not in ['필수이수', '선택이수']: 
                continue
            
            courses_list.append({
                "grade_sem": str(row.get(sem_col, '')).strip(),
                "course_name": str(row.get(name_col, '')).strip(),
                "area_1": c_area
            })

        # 2. 지정된 규칙에 따른 정렬 (필수이수 18개 우선 순위 적용)
        def sort_key(item):
            area = item["area_1"]
            c_name = item["course_name"].replace(" ", "")
            g_sem = item["grade_sem"]
            area_order = 0 if area == '필수이수' else 1
            custom_idx = 999
            if area_order == 0:
                if "보건의료정보관리학" in c_name: custom_idx = 1
                elif "보건의료정보관리실무" in c_name: custom_idx = 2
                elif "보건의료조직" in c_name: custom_idx = 3
                elif "건강정보보호" in c_name: custom_idx = 4
                elif "질병" in c_name and "분류" in c_name: custom_idx = 5
                elif "의무기록정보분석" in c_name: custom_idx = 6
                elif "의무기록정보질" in c_name: custom_idx = 7
                elif "암등록" in c_name: custom_idx = 8
                elif "의료의질" in c_name or "의료질" in c_name: custom_idx = 9
                elif "건강보험" in c_name: custom_idx = 10
                elif "보건의료통계" in c_name: custom_idx = 11
                elif "보건의료데이터" in c_name: custom_idx = 12
                elif "의료정보기술" in c_name: custom_idx = 13
                elif "의료관계법규" in c_name: custom_idx = 14
                elif "의학용어" in c_name: custom_idx = 15
                elif "병리학" in c_name: custom_idx = 16
                elif "해부생리" in c_name: custom_idx = 17
                elif "현장실습" in c_name: custom_idx = 18
            return (area_order, custom_idx, g_sem, item["course_name"])
            
        courses_list.sort(key=sort_key)

        # 3. 엑셀 워크북 생성 및 공통 스타일 설정
        wb = Workbook()
        ws = wb.active
        ws.title = f"{year}학년도 전체명세서"
        
        grey_bg = PatternFill("solid", fgColor="D9D9D9")
        white_bg = PatternFill("solid", fgColor="FFFFFF")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        def set_cell(r, c, val, fill=None, font_color="000000", bold=False, align='center'):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
            if fill: cell.fill = fill
            cell.font = Font(color=font_color, bold=bold)
            cell.border = thin_border
            return cell

        # 4. 각 과목별 명세서 작성 루프
        current_row = 1
        for idx, c_info in enumerate(courses_list):
            cname = c_info["course_name"]
            data = {"year": year, "course_name": cname, "area_1": c_info["area_1"]}
            
            # 저장된 파일(Draft 우선, 없으면 최신 Version) 탐색 및 로드
            s_dir = os.path.join(UPLOAD_DIR, "명세서", str(year))
            target = os.path.join(s_dir, f"{cname}_draft.csv")
            if not os.path.exists(target) and os.path.exists(s_dir):
                files = [f for f in os.listdir(s_dir) if f.startswith('v') and cname in f and f.endswith('.csv')]
                if files:
                    files.sort(reverse=True)
                    target = os.path.join(s_dir, files[0])
            
            df_saved = smart_read_df(target)
            if df_saved is not None and not df_saved.empty:
                data.update(df_saved.fillna("").iloc[0].to_dict())

            # --- [엑셀 작성 시작] ---
            r = current_row
            
            # (1) 제목 행
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
            set_cell(r, 1, f"{idx+1}. {cname} (입학년도: {year})", None, bold=True, align='left')
            r += 1
            
            # (2) 기본 정보
            ws.merge_cells(f"A{r}:B{r}"); set_cell(r, 1, "교과목명", grey_bg, bold=True)
            ws.merge_cells(f"C{r}:D{r}"); set_cell(r, 3, cname)
            ws.merge_cells(f"E{r}:F{r}"); set_cell(r, 5, "구분", grey_bg, bold=True)
            ws.merge_cells(f"G{r}:H{r}"); set_cell(r, 7, data.get('area_1', ''))
            ws.merge_cells(f"I{r}:J{r}"); set_cell(r, 9, "책임교수", grey_bg, bold=True)
            ws.merge_cells(f"K{r}:L{r}"); set_cell(r, 11, data.get('professor', ''))
            r += 1
            
            # (3) 교과목개요
            ws.merge_cells(f"A{r}:B{r}"); set_cell(r, 1, "교과목개요", grey_bg, bold=True)
            ws.merge_cells(f"C{r}:L{r}"); set_cell(r, 3, data.get('overview', ''), align='left')
            r += 1
            
            # (4) 프로그램 최종성과(PO)
            ws.merge_cells(f"A{r}:B{r+1}"); set_cell(r, 1, "프로그램\n최종성과", grey_bg, bold=True)
            for i in range(10): set_cell(r, i+3, f"PO{i+1}", grey_bg, bold=True)
            r += 1
            for i in range(10): set_cell(r, i+3, data.get(f'po{i+1}', ''))
            r += 1
            
            # (5) 학습목표 (CLO)
            try: clo_list = json.loads(data.get('clo_list', '[]'))
            except: clo_list = []
            
            clo_rows = max(len(clo_list), 1)
            ws.merge_cells(start_row=r, start_column=1, end_row=r + clo_rows, end_column=2)
            set_cell(r, 1, "학습목표\n(교과목 학습성과)", grey_bg, bold=True)
            
            ws.merge_cells(f"C{r}:H{r}"); set_cell(r, 3, "학습목표 (CLO)", grey_bg, bold=True)
            ws.merge_cells(f"I{r}:J{r}"); set_cell(r, 9, "연계 PO", grey_bg, bold=True)
            ws.merge_cells(f"K{r}:L{r}"); set_cell(r, 11, "PO 평가반영", grey_bg, bold=True)
            r += 1
            
            if not clo_list:
                ws.merge_cells(f"C{r}:H{r}"); set_cell(r, 3, ""); set_cell(r, 9, ""); set_cell(r, 11, "")
                r += 1
            else:
                for clo in clo_list:
                    set_cell(r, 3, clo.get('no', ''), grey_bg, bold=True)
                    ws.merge_cells(f"D{r}:H{r}"); set_cell(r, 4, clo.get('content', ''), align='left')
                    ws.merge_cells(f"I{r}:J{r}"); set_cell(r, 9, data.get('linked_po',''))
                    ws.merge_cells(f"K{r}:L{r}"); set_cell(r, 11, data.get('po_reflected',''))
                    r += 1
            
            # (6) 학습내용 및 교수법
            ws.merge_cells(f"A{r}:B{r}"); set_cell(r, 1, "학습내용", grey_bg, bold=True)
            ws.merge_cells(f"C{r}:L{r}"); set_cell(r, 3, data.get('learning_content', ''), align='left')
            r += 1
            ws.merge_cells(f"A{r}:B{r}"); set_cell(r, 1, "교수법", grey_bg, bold=True)
            ws.merge_cells(f"C{r}:L{r}"); set_cell(r, 3, data.get('teaching_method', ''), align='left')
            r += 1
            
            # (7) 평가 기준 (CLO별 배점)
            try: eval_data = json.loads(data.get('eval_data', '[]'))
            except: eval_data = []
            ecols = len(eval_data) if eval_data else 1
            
            ws.merge_cells(f"A{r}:A{r+4}"); set_cell(r, 1, "평가", grey_bg, bold=True)
            ws.merge_cells(f"B{r}:B{r+1}"); set_cell(r, 2, "영역", grey_bg, bold=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=12)
            set_cell(r, 3, "학습목표 (CLO) 성취도", grey_bg, bold=True)
            r += 1
            
            # CLO 헤더 및 데이터 채우기 (열 너비 균등 분할)
            c_start = 3
            for i in range(ecols):
                span = (10 // ecols) + (1 if i < (10 % ecols) else 0)
                c_end = c_start + span - 1
                if span > 1: ws.merge_cells(start_row=r, start_column=c_start, end_row=r, end_column=c_end)
                set_cell(r, c_start, f"CLO{i+1}", grey_bg, bold=True)
                # 배점율
                set_cell(r+1, c_start, eval_data[i].get('criteria','') if eval_data else "", white_bg)
                if span > 1: ws.merge_cells(start_row=r+1, start_column=c_start, end_row=r+1, end_column=c_end)
                # 평가방법
                set_cell(r+2, c_start, eval_data[i].get('method','') if eval_data else "", white_bg)
                if span > 1: ws.merge_cells(start_row=r+2, start_column=c_start, end_row=r+2, end_column=c_end)
                # 목표수준
                set_cell(r+3, c_start, eval_data[i].get('target','') if eval_data else "", white_bg)
                if span > 1: ws.merge_cells(start_row=r+3, start_column=c_start, end_row=r+3, end_column=c_end)
                c_start = c_end + 1
            
            set_cell(r+1, 2, "배점율", grey_bg, bold=True)
            set_cell(r+2, 2, "평가방법", grey_bg, bold=True)
            set_cell(r+3, 2, "목표수준", grey_bg, bold=True)
            r += 4
            
            # (8) 선수과목
            ws.merge_cells(f"A{r}:B{r}"); set_cell(r, 1, "선수과목", grey_bg, bold=True)
            ws.merge_cells(f"C{r}:L{r}"); set_cell(r, 3, data.get('prerequisite', ''), align='left')
            
            # 과목 간 간격 (3줄 띄우기)
            current_row = r + 4

        # 5. 파일 스트리밍 반환
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fname = urllib.parse.quote(f"{year}학년도_명세서_통합본.xlsx")
        return StreamingResponse(output, 
                                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"})
                                
    except Exception as e:
        print(f"❌ 전체 엑셀 생성 중 에러 발생: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/명세서/status")
def get_myungseseo_status(year: str):
    s_dir = os.path.join(UPLOAD_DIR, "명세서", str(year))
    if not os.path.exists(s_dir):
        return {"completed_courses": []}
    
    completed = []
    for f in os.listdir(s_dir):
        # 'v'로 시작하고 '.csv'로 끝나는 파일이 최종 저장 파일입니다. (예: v1_과목명_20260309.csv)
        if f.startswith('v') and f.endswith('.csv'):
            parts = f.split('_')
            if len(parts) >= 3:
                c_name = "_".join(parts[1:-1])  # 과목명만 정확히 추출
                completed.append(c_name)
                
    return {"completed_courses": list(set(completed))}

# ==========================================
# ★★★ [규정집 DB 전용 API] ★★★
# ==========================================
RULEBOOK_CSV_PATH = os.path.join(UPLOAD_DIR, "규정집_DB.csv")

@app.get("/api/규정집/list")
def get_rulebook_list():
    if not os.path.exists(RULEBOOK_CSV_PATH):
        return {"rules": []}
    try:
        df = pd.read_csv(RULEBOOK_CSV_PATH, encoding='utf-8-sig')
        df = df.where(pd.notnull(df), "")
        return {"rules": df.to_dict(orient="records")}
    except Exception as e:
        print("규정집 로드 에러:", e)
        return {"rules": []}

@app.post("/api/규정집/save")
def save_rulebook(payload: dict = Body(...)):
    # payload: { rule_id, url, rule_name, revision_date, task_date, content }
    if os.path.exists(RULEBOOK_CSV_PATH):
        df = pd.read_csv(RULEBOOK_CSV_PATH, encoding='utf-8-sig')
    else:
        df = pd.DataFrame(columns=["rule_id", "url", "rule_name", "revision_date", "task_date", "content"])
    
    # 이미 존재하는 ID면 업데이트, 아니면 추가
    idx = df.index[df['rule_id'] == payload.get('rule_id')].tolist()
    if idx:
        for k, v in payload.items():
            df.at[idx[0], k] = v
    else:
        df = pd.concat([df, pd.DataFrame([payload])], ignore_index=True)
        
    df.to_csv(RULEBOOK_CSV_PATH, index=False, encoding='utf-8-sig')
    return {"status": "success"}

@app.post("/api/규정집/crawl")
def crawl_yonsei_rules(payload: dict = Body(...)):
    target_url = payload.get("url", "")
    try:
        if "search#" in target_url:
            rule_id = target_url.split("search#")[1].strip()
            print_url = f"https://rules.yonsei.ac.kr/ruleseq{rule_id}/print"
        else:
            print_url = target_url

        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(print_url, headers=headers, timeout=10)
        response.encoding = 'utf-8' 
        soup = BeautifulSoup(response.text, 'html.parser')

        # 1. 규정명 추출
        rule_name = "규정명을 찾을 수 없습니다"
        title_tag = soup.find('title')
        if title_tag:
            rule_name = title_tag.text.split('::')[0].strip()
            
        if rule_name == "규정명을 찾을 수 없습니다" or not rule_name:
            hx = soup.find(['h1', 'h2', 'h3'])
            if hx: rule_name = hx.text.strip()

        # 2. 본문, 개정일, ★담당부서 추출
        all_text = soup.get_text(separator='\n')
        lines = all_text.split('\n')
        
        revision_date = "날짜 없음"
        department = "부서 정보 없음" # ★ 추가됨
        content_lines = []
        is_content_started = False
        
        for line in lines:
            line_stripped = line.strip()
            if not line_stripped:
                continue
                
            if "개정일" in line_stripped or re.search(r'\d{4}\.\d{2}\.\d{2}', line_stripped):
                if revision_date == "날짜 없음":
                    match = re.search(r'\d{4}\.\d{2}\.\d{2}', line_stripped)
                    if match: revision_date = match.group()
                    continue

            # ★ 담당부서 찾기
            if "담당부서" in line_stripped:
                if department == "부서 정보 없음":
                    department = line_stripped.replace("담당부서:", "").replace("담당부서", "").strip()
                continue
            
            # 본문 시작점 찾기
            if re.match(r'^제\s*\d+\s*[조장]', line_stripped):
                is_content_started = True
                
            if is_content_started:
                content_lines.append(line_stripped)

        content_text = "\n\n".join(content_lines)
        if not content_text:
            content_text = "본문을 분리할 수 없습니다.\n\n" + all_text[:1500]

        return {
            "rule_name": rule_name,
            "revision_date": revision_date,
            "department": department, # ★ 추가됨
            "content": content_text
        }
    except Exception as e:
        print("크롤링 에러 상세:", str(e))
        raise HTTPException(status_code=500, detail="크롤링 중 오류가 발생했습니다.")

@app.post("/api/규정집/download/{format}")
def download_rulebook(format: str, payload: dict = Body(...)):
    rules = payload.get("rules", [])
    if not rules:
        raise HTTPException(status_code=400, detail="다운로드할 데이터가 없습니다.")
    
    # 1. 단일 문서(Document) 생성
    doc = Document()
    
    # 문서 전체 기본 폰트 설정 (맑은 고딕)
    style = doc.styles['Normal']
    style.font.name = '맑은 고딕'
    style._element.rPr.rFonts.set(qn('w:eastAsia'), '맑은 고딕')
    style.font.size = Pt(11)

    # 2. 선택된 규정들을 순서대로 하나의 문서에 이어 붙이기
    for i, rule_data in enumerate(rules):
        # 1) 문서 제목
        p_title = doc.add_paragraph()
        p_title.alignment = WD_ALIGN_PARAGRAPH.CENTER
        run_title = p_title.add_run(rule_data.get("rule_name", "제목없음"))
        run_title.bold = True
        run_title.font.size = Pt(22)
        p_title.paragraph_format.space_after = Pt(20)
        
        # 2) 메타 데이터
        p_meta = doc.add_paragraph()
        p_meta.alignment = WD_ALIGN_PARAGRAPH.RIGHT
        meta_text = f"개정일: {rule_data.get('revision_date', '')}\n담당부서: {rule_data.get('department', '')}\n작업 날짜: {rule_data.get('task_date', '')}"
        run_meta = p_meta.add_run(meta_text)
        run_meta.font.size = Pt(10)
        run_meta.font.color.rgb = RGBColor(80, 80, 80)
        
        # 구분선
        p_sep = doc.add_paragraph("―" * 45)
        p_sep.alignment = WD_ALIGN_PARAGRAPH.CENTER
        p_sep.paragraph_format.space_after = Pt(20)

        # 3) 본문 내용 포맷팅
        content = rule_data.get("content", "")
        lines = content.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue

            p = doc.add_paragraph()
            
            if re.match(r'^제\s*\d+\s*[장절]', line):
                p.alignment = WD_ALIGN_PARAGRAPH.CENTER
                run = p.add_run(line)
                run.bold = True
                run.font.size = Pt(15) if "장" in line else Pt(13)
                p.paragraph_format.space_before = Pt(20)
                p.paragraph_format.space_after = Pt(10)
                
            elif re.match(r'^(제\s*\d+\s*조(?:의\d+)?\s*(?:\([^)]+\))?)(.*)', line):
                article_match = re.match(r'^(제\s*\d+\s*조(?:의\d+)?\s*(?:\([^)]+\))?)(.*)', line)
                run_bold = p.add_run(article_match.group(1))
                run_bold.bold = True
                run_bold.font.size = Pt(12)
                if article_match.group(2):
                    run_text = p.add_run(" " + article_match.group(2).strip())
                    run_text.font.size = Pt(11)
                p.paragraph_format.space_before = Pt(15)
                
            else:
                if re.match(r'^[①②③④⑤⑥⑦⑧⑨⑩\d+\.]', line):
                    p.paragraph_format.left_indent = Pt(15)
                run = p.add_run(line)
                run.font.size = Pt(11)

        # ★ 4) 규정이 더 남아있다면 다음 페이지로 넘기기 (Page Break)
        if i < len(rules) - 1:
            doc.add_page_break()

    # 3. 완성된 단일 문서를 저장 및 반환
    doc_io = BytesIO()
    doc.save(doc_io)
    doc_io.seek(0)
    
    # 다운로드 파일명 결정 (1개면 해당 규정명, 여러 개면 '통합문서')
    if len(rules) == 1:
        r_name = rules[0].get("rule_name", "규정명없음").replace("/", "_")
        encoded_filename = urllib.parse.quote(f"{r_name}.docx")
    else:
        encoded_filename = urllib.parse.quote("규정집_통합다운로드.docx")
    
    return StreamingResponse(
        doc_io, 
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document", 
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )

@app.post("/api/규정집/delete")
def delete_rulebook(payload: dict = Body(...)):
    rule_ids = payload.get("rule_ids", [])
    if not rule_ids:
        return {"status": "success"}

    if os.path.exists(RULEBOOK_CSV_PATH):
        df = pd.read_csv(RULEBOOK_CSV_PATH, encoding='utf-8-sig')
        # 전달받은 rule_id 목록에 포함되지 않은 데이터만 남기기 (삭제 효과)
        df = df[~df['rule_id'].isin(rule_ids)]
        df.to_csv(RULEBOOK_CSV_PATH, index=False, encoding='utf-8-sig')
        
    return {"status": "success"}

# ==========================================
# ★★★ [강의계획서 DB 전용 API (연도별 분리 & 원본 PDF 뷰어)] ★★★
# ==========================================
@app.get("/api/강의계획서/status")
def get_syllabus_status(year: str):
    """특정 학년도에 업로드된 강의계획서 목록 반환"""
    s_dir = os.path.join(UPLOAD_DIR, "강의계획서", str(year))
    if not os.path.exists(s_dir):
        return {"uploaded_courses": []}
    files = os.listdir(s_dir)
    courses = [f.replace('.pdf', '') for f in files if f.endswith('.pdf')]
    return {"uploaded_courses": courses}

@app.post("/api/강의계획서/upload")
async def upload_syllabus(year: str = Form(...), course_name: str = Form(...), file: UploadFile = File(...)):
    """특정 학년도 폴더에 PDF 원본 업로드"""
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="PDF 파일만 업로드 가능합니다.")
        
    s_dir = os.path.join(UPLOAD_DIR, "강의계획서", str(year))
    os.makedirs(s_dir, exist_ok=True)
    
    file_path = os.path.join(s_dir, f"{course_name}.pdf")
    
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
        
    return {"status": "success", "message": "업로드 완료"}

@app.get("/api/강의계획서/view")
def view_syllabus(year: str, course_name: str, t: str = None):
    safe_name = os.path.basename(course_name.replace("..", ""))
    file_path = os.path.join(UPLOAD_DIR, "강의계획서", str(year), f"{safe_name}.pdf")
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="파일이 존재하지 않습니다.")
        
    encoded_name = urllib.parse.quote(f"{safe_name}.pdf")
    return FileResponse(
        file_path, 
        media_type="application/pdf", 
        headers={"Content-Disposition": f"inline; filename*=UTF-8''{encoded_name}"}
    )

@app.get("/api/강의계획서/download-all-pdf")
def download_all_syllabus_pdf(year: str):
    """명세서처럼 학년도 전체 과목을 순서대로 하나의 PDF로 병합하여 다운로드"""
    # 1. 교과목 DB에서 해당 학년도의 '정렬된 순서'를 가져옴
    courses_res = get_myungseseo_courses(year)
    courses_list = courses_res.get("courses", [])
    
    merger = PdfWriter()
    merged_count = 0
    s_dir = os.path.join(UPLOAD_DIR, "강의계획서", str(year))
    
    # 2. 정렬된 순서대로 PDF가 있으면 이어붙임
    for c in courses_list:
        c_name = c["course_name"]
        file_path = os.path.join(s_dir, f"{c_name}.pdf")
        if os.path.exists(file_path):
            merger.append(file_path)
            merged_count += 1
            
    if merged_count == 0:
        raise HTTPException(status_code=404, detail="해당 학년도에 업로드된 강의계획서가 없습니다.")

    out_io = BytesIO()
    merger.write(out_io)
    merger.close()
    out_io.seek(0)
    
    encoded_filename = urllib.parse.quote(f"{year}학년도_전체강의계획서.pdf")
    return StreamingResponse(
        out_io, media_type="application/pdf", 
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )

# ==========================================
# 교수진 강의담당 분석 관련 API
# ==========================================
def extract_rooms_from_analysis(analysis_df: pd.DataFrame):
    # 'area_1' 컬럼(교과영역)에서 필수이수, 선택이수 필터링
    target_areas = ['필수이수', '선택이수']
    if 'area_1' in analysis_df.columns:
        filtered_df = analysis_df[analysis_df['area_1'].isin(target_areas)]
    elif '교과영역' in analysis_df.columns:
        filtered_df = analysis_df[analysis_df['교과영역'].isin(target_areas)]
    else:
        filtered_df = analysis_df  # 해당 컬럼이 없으면 전체 데이터로 진행

    building_map = {
        '미': '미래관',
        '창': '창조관',
        '컨': '컨버전스',
        '청': '청송관',
        '백': '백운관'
    }

    room_data = []
    has_other_building = False

    for _, row in filtered_df.iterrows():
        # 학기 분리 (예: '2026-1학기' -> 2026, 1학기)
        semester_str = str(row.get('학기', '')).strip()
        match = re.search(r'(\d{4})-(.*)', semester_str)
        if match:
            room_year = match.group(1)
            room_semester = match.group(2)
        else:
            room_year = ''
            room_semester = semester_str

        # 건물1, 호실1 추출
        b1 = str(row.get('건물1', '')).strip()
        r1 = str(row.get('호실1', '')).strip()
        
        if b1 and b1 != 'nan':
            mapped_b1 = building_map.get(b1, '기타')
            if mapped_b1 == '기타':
                has_other_building = True
            
            room_data.append({
                'room_year': room_year,
                'room_semester': room_semester,
                'room_building': mapped_b1,
                'room_number': r1,
                'room_usage_type': '공용',
                'room_capacity': '', # 자동완성용 빈칸
                'room_area': ''      # 자동완성용 빈칸
            })

        # 건물2, 호실2 추출 (독립적인 데이터로 분리 추가)
        b2 = str(row.get('건물2', '')).strip()
        r2 = str(row.get('호실2', '')).strip()
        
        if b2 and b2 != 'nan':
            mapped_b2 = building_map.get(b2, '기타')
            if mapped_b2 == '기타':
                has_other_building = True
                
            room_data.append({
                'room_year': room_year,
                'room_semester': room_semester,
                'room_building': mapped_b2,
                'room_number': r2,
                'room_usage_type': '공용',
                'room_capacity': '',
                'room_area': ''
            })

    # 데이터프레임 변환
    result_df = pd.DataFrame(room_data, columns=[
        'room_year', 'room_semester', 'room_building', 'room_number', 
        'room_usage_type', 'room_capacity', 'room_area'
    ])
    
    # 완전히 동일한 강의실 중복 제거 (학년도, 학기, 건물, 호수 기준)
    if not result_df.empty:
        result_df = result_df.drop_duplicates(
            subset=['room_year', 'room_semester', 'room_building', 'room_number']
        ).reset_index(drop=True)

    return result_df, has_other_building

# 2. 프론트엔드에서 강의실 데이터를 요청할 때 호출되는 API 엔드포인트
@app.get("/api/extract-rooms")
def api_extract_rooms():
    # UPLOAD_DIR은 main.py 상단에 정의된 변수 (기본값 'uploads')
    upload_path = globals().get('UPLOAD_DIR', 'uploads')
    target_dir = os.path.join(upload_path, "교수진 강의담당 분석")
    
    if not os.path.exists(target_dir):
        return {"status": "error", "message": "'교수진 강의담당 분석' 데이터가 존재하지 않습니다. 먼저 업로드해 주세요."}
        
    room_data = []
    has_other_building = False
    
    # 건물명 매핑 사전
    building_map = {
        '미': '미래관', '창': '창조관', '컨': '컨버전스', '청': '청송관', '백': '백운관'
    }
    
    # 추출할 교과영역 (필요시 추가 가능)
    target_areas = ['필수이수', '선택이수', '전공선택', '전공필수']

    for root, _, files in os.walk(target_dir):
        for file in files:
            if file.lower().endswith(('.csv', '.xlsx')):
                file_path = os.path.join(root, file)
                
                # 1. 파일명에서 202X 학년도 추출 (예: 2026_draft.csv -> 2026)
                year_match = re.search(r'(202\d)', file)
                file_year = year_match.group(1) if year_match else ''
                
                try:
                    # 2. 파일 읽기 (인코딩 에러 방지)
                    if file.lower().endswith('.csv'):
                        try:
                            df = pd.read_csv(file_path, encoding='utf-8-sig')
                        except UnicodeDecodeError:
                            df = pd.read_csv(file_path, encoding='cp949')
                    else:
                        df = pd.read_excel(file_path, engine='openpyxl')
                        
                    # 3. 영문/한글 컬럼명 유연하게 찾기
                    area_col = next((c for c in ['area_1', '교과영역'] if c in df.columns), None)
                    sem_col = next((c for c in ['semester', '학기'] if c in df.columns), None)
                    b1_col = next((c for c in ['first_build', '건물1'] if c in df.columns), None)
                    r1_col = next((c for c in ['first_room', '호실1'] if c in df.columns), None)
                    b2_col = next((c for c in ['second_build', '건물2'] if c in df.columns), None)
                    r2_col = next((c for c in ['second_room', '호실2'] if c in df.columns), None)

                    # 교과영역 필터링
                    if area_col:
                        filtered_df = df[df[area_col].isin(target_areas)]
                    else:
                        filtered_df = df

                    for _, row in filtered_df.iterrows():
                        semester_str = str(row.get(sem_col, '')).strip() if sem_col else ''
                        
                        # 학기 분리 (예: '2026-1학기' -> 2026, 1학기 / 그냥 '1학기' -> 파일명연도, 1학기)
                        match = re.search(r'(\d{4})-(.*)', semester_str)
                        if match:
                            room_year = match.group(1)
                            room_semester = match.group(2)
                        else:
                            room_year = file_year
                            room_semester = semester_str
                            
                        # 건물1, 호실1 추출
                        b1 = str(row.get(b1_col, '')).strip() if b1_col else ''
                        r1 = str(row.get(r1_col, '')).strip() if r1_col else ''
                        if r1.endswith('.0'): r1 = r1[:-2] 
                        elif '.' in r1: r1 = r1.split('.')[0] 
                        
                        if b1 and b1 != 'nan':
                            mapped_b1 = building_map.get(b1, b1) # 맵에 없으면 엑셀에 적힌 원본 사용
                            room_data.append({
                                'room_year': room_year, 'room_semester': room_semester,
                                'room_building': mapped_b1, 'room_number': r1,
                                'room_usage_type': '공용', 'room_capacity': '', 'room_area': ''
                            })
                            
                        # 건물2, 호실2 추출
                        b2 = str(row.get(b2_col, '')).strip() if b2_col else ''
                        r2 = str(row.get(r2_col, '')).strip() if r2_col else ''
                        if r2.endswith('.0'): r2 = r2[:-2] 
                        elif '.' in r2: r2 = r2.split('.')[0] 
                        
                        if b2 and b2 != 'nan':
                            mapped_b2 = building_map.get(b2, b2)
                            room_data.append({
                                'room_year': room_year, 'room_semester': room_semester,
                                'room_building': mapped_b2, 'room_number': r2,
                                'room_usage_type': '공용', 'room_capacity': '', 'room_area': ''
                            })

                except Exception as e:
                    print(f"강의실 추출용 파일 읽기 에러 {file}: {e}")
    
    # 4. 데이터프레임 변환 및 중복 제거
    result_df = pd.DataFrame(room_data, columns=[
        'room_year', 'room_semester', 'room_building', 'room_number', 
        'room_usage_type', 'room_capacity', 'room_area'
    ])
    
    if not result_df.empty:
        result_df = result_df.drop_duplicates(
            subset=['room_year', 'room_semester', 'room_building', 'room_number']
        ).reset_index(drop=True)
        
    return {
        "status": "success",
        "has_other_building": has_other_building,
        "rows": result_df.fillna("").to_dict(orient='records'),
        "count": len(result_df)
    }

@app.get("/api/auto-complete-rooms")
def auto_complete_rooms(current_year: str, current_semester: str):
    # 직전 학기 계산 (1학기, 겨울학기 -> 직전년도 2학기 / 2학기, 여름학기 -> 해당년도 1학기)
    try:
        year_int = int(current_year)
    except ValueError:
        return {"status": "error", "message": "유효하지 않은 학년도입니다."}

    if current_semester in ["1학기", "겨울학기"]:
        prev_year = str(year_int - 1)
        prev_semester = "2학기"
    else:
        prev_year = str(year_int)
        prev_semester = "1학기"
        
    df, error = get_db_dataframe("강의실")
    if error or df is None or df.empty:
        return {"status": "success", "data": {}}
        
    # 직전 학기 데이터만 필터링
    prev_df = df[(df['room_year'].astype(str) == prev_year) & (df['room_semester'].astype(str) == prev_semester)]
    
    # 건물+호수를 Key로 하여 수용인원과 면적 데이터를 딕셔너리로 생성
    auto_complete_data = {}
    for _, row in prev_df.iterrows():
        # 호실명 소수점 제거
        r_num = str(row.get('room_number', '')).strip()
        if r_num.endswith('.0'): r_num = r_num[:-2]
        
        key = f"{row.get('room_building', '')}_{r_num}"
        
        # 수용인원 소수점 제거
        cap = str(row.get('room_capacity', '')).strip()
        if cap.endswith('.0'): cap = cap[:-2]
        
        auto_complete_data[key] = {
            'room_capacity': cap,
            'room_area': row.get('room_area', '')
        }
        
    return {
        "status": "success", 
        "data": auto_complete_data, 
        "prev_term": f"{prev_year}학년도 {prev_semester}"
    }

# 엑셀 다운로드를 위해 프론트엔드에서 보내는 데이터 구조 정의
class ExcelRequest(BaseModel):
    data: List[Dict[str, Any]]

@app.post("/api/download/excel/{db_name}")
async def download_excel(db_name: str, request: ExcelRequest):
    try:
        # 1. 프론트엔드에서 받은 데이터를 판다스(Pandas) 표로 변환
        df = pd.DataFrame(request.data)
        
        # 2. 가상의 메모리 공간에 엑셀 파일 생성
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            safe_sheet_name = db_name[:31] # 엑셀 시트 이름은 31자 제한
            df.to_excel(writer, index=False, sheet_name=safe_sheet_name)
        
        # 3. 한글 파일명이 깨지지 않도록 변환
        encoded_filename = urllib.parse.quote(f"{db_name}.xlsx")
        
        # 4. 완성된 엑셀 파일을 브라우저로 전송
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    except Exception as e:
        print(f"엑셀 다운로드 오류: {e}")
        return {"status": "error", "message": str(e)}

# ==========================================
# ★★★ [캘린더 및 증빙 서류 관리 전용 API] ★★★
# ==========================================
CALENDAR_DB_PATH = os.path.join(UPLOAD_DIR, "calendar_events.json")

@app.post("/api/calendar/events")
def save_calendar_event(payload: dict = Body(...)):
    events = []
    if os.path.exists(CALENDAR_DB_PATH):
        with open(CALENDAR_DB_PATH, "r", encoding="utf-8") as f:
            try: events = json.load(f)
            except: pass
    
    # ⭐️ 기존에 같은 날짜(start)가 있으면 삭제하고 새로 넣음 (업데이트 로직)
    events = [e for e in events if e.get('start') != payload.get('start')]
    
    events.append(payload)
    with open(CALENDAR_DB_PATH, "w", encoding="utf-8") as f:
        json.dump(events, f, ensure_ascii=False, indent=4)
    return {"status": "success"}

@app.get("/api/calendar/events")
def get_calendar_events(year: str = None):
    if not os.path.exists(CALENDAR_DB_PATH):
        return []
    with open(CALENDAR_DB_PATH, "r", encoding="utf-8") as f:
        events = json.load(f)
    if year:
        # 연도별 필터링
        return [e for e in events if e.get('year') == year]
    return events

@app.post("/api/calendar/events/update")
def update_calendar_event(payload: dict = Body(...)):
    # 특정 날짜와 내용이 일치하는 이벤트를 찾아 업데이트
    events = []
    if os.path.exists(CALENDAR_DB_PATH):
        with open(CALENDAR_DB_PATH, "r", encoding="utf-8") as f:
            events = json.load(f)
    
    for i, e in enumerate(events):
        if e.get('date') == payload.get('date') and e.get('content') == payload.get('content'):
            events[i] = payload
            break
            
    with open(CALENDAR_DB_PATH, "w", encoding="utf-8") as f:
        json.dump(events, f, ensure_ascii=False, indent=4)
    return {"status": "success"}

@app.post("/api/calendar/upload")
def upload_calendar_file(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg')):
        raise HTTPException(status_code=400, detail="이미지 파일(PNG, JPG)만 업로드 가능합니다.")
        
    upload_path = os.path.join(UPLOAD_DIR, "calendar")
    os.makedirs(upload_path, exist_ok=True)
    
    # 안전한 파일명 생성
    safe_filename = f"{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
    file_location = os.path.join(upload_path, safe_filename)
    
    with open(file_location, "wb+") as f:
        shutil.copyfileobj(file.file, f)
        
    # 프론트엔드에서 이미지를 불러올 수 있는 URL 경로 반환
    return {"file_url": f"http://localhost:8000/uploads/calendar/{safe_filename}", "file_name": file.filename}

# main.py

@app.post("/api/calendar/save-all")
def save_calendar_all(payload: dict = Body(...)):
    year = payload.get("year")
    tasks = payload.get("tasks")
    save_type = payload.get("type") # 'temp' 또는 'final'

    # 기존 데이터 로드
    all_events = []
    if os.path.exists(CALENDAR_DB_PATH):
        with open(CALENDAR_DB_PATH, "r", encoding="utf-8") as f:
            all_events = json.load(f)

    # 해당 연도의 데이터만 새 데이터로 교체
    other_years_events = [e for e in all_events if e.get('year') != year]
    new_all_events = other_years_events + tasks

    with open(CALENDAR_DB_PATH, "w", encoding="utf-8") as f:
        json.dump(new_all_events, f, ensure_ascii=False, indent=4)
        
    return {"status": "success"}
  

# ============================================================
#  설문 엔드포인트 — main.py의 if __name__ == "__main__": 바로 위에 붙여넣기
# ============================================================
import os, json, datetime
from fastapi import FastAPI, HTTPException
from typing import Any, Dict

SURVEY_DIR = os.path.join(os.path.dirname(__file__), "surveys")
os.makedirs(SURVEY_DIR, exist_ok=True)

VALID_KEYS = {"s0", "s1", "s2", "s3", "s4", "s5", "s6"}

def _survey_path(survey_key: str) -> str:
    return os.path.join(SURVEY_DIR, f"{survey_key}.json")

def _load(survey_key: str) -> list:
    path = _survey_path(survey_key)
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def _save(survey_key: str, data: list):
    with open(_survey_path(survey_key), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


@app.post("/api/surveys/{survey_key}")
def submit_survey(survey_key: str, body: Dict[str, Any] = Body(...)):
    """설문 응답 제출"""
    if survey_key not in VALID_KEYS:
        raise HTTPException(status_code=404, detail="존재하지 않는 설문입니다.")
    responses = _load(survey_key)
    body["_submitted_at"] = datetime.datetime.now().isoformat()
    responses.append(body)
    _save(survey_key, responses)
    return {"ok": True, "count": len(responses)}


@app.get("/api/surveys/{survey_key}/results")
def get_survey_results(survey_key: str):
    """설문 결과 조회"""
    if survey_key not in VALID_KEYS:
        raise HTTPException(status_code=404, detail="존재하지 않는 설문입니다.")
    responses = _load(survey_key)
    return {"survey_key": survey_key, "count": len(responses), "responses": responses}


@app.delete("/api/surveys/{survey_key}")
def clear_survey(survey_key: str):
    """설문 응답 전체 삭제 (관리용)"""
    if survey_key not in VALID_KEYS:
        raise HTTPException(status_code=404, detail="존재하지 않는 설문입니다.")
    _save(survey_key, [])
    return {"ok": True, "message": f"{survey_key} 응답이 초기화되었습니다."}

# ── 설문 엔드포인트 ──────────────────────────────────────
import os, json, datetime
from fastapi import Body
from typing import Any, Dict

SURVEY_DIR = os.path.join(os.path.dirname(__file__), "surveys")
os.makedirs(SURVEY_DIR, exist_ok=True)
VALID_KEYS = {"s0", "s1", "s2", "s3", "s4", "s5", "s6"}

def _survey_path(k): return os.path.join(SURVEY_DIR, f"{k}.json")
def _load(k):
    p = _survey_path(k)
    return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else []
def _save(k, data):
    json.dump(data, open(_survey_path(k), "w", encoding="utf-8"), ensure_ascii=False, indent=2)

@app.post("/api/surveys/{survey_key}")
def submit_survey(survey_key: str, body: Dict[str, Any] = Body(...)):
    if survey_key not in VALID_KEYS: raise HTTPException(status_code=404, detail="없는 설문")
    rows = _load(survey_key)
    body["_submitted_at"] = datetime.datetime.now().isoformat()
    rows.append(body)
    _save(survey_key, rows)
    return {"ok": True, "count": len(rows)}

@app.get("/api/surveys/{survey_key}/results")
def get_survey_results(survey_key: str):
    if survey_key not in VALID_KEYS: raise HTTPException(status_code=404, detail="없는 설문")
    rows = _load(survey_key)
    return {"survey_key": survey_key, "count": len(rows), "responses": rows}

if __name__ == "__main__": 
    uvicorn.run(app, host="0.0.0.0", port=8000)